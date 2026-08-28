import openpyxl, json, os, math
R='/home/claude/repo/'
wb=openpyxl.load_workbook('/home/claude/chp.xlsx', read_only=True, data_only=True)

# --- metadata sheet gives the human name, definition, source and years ---
mws=wb['Metadata']
mrows=[r for r in mws.iter_rows(values_only=True)]
mhdr=[str(c) if c else '' for c in mrows[0]]
META={}
for r in mrows[1:]:
    if not r[1]: continue
    for v in str(r[1]).split(','):
        v=v.strip()
        if v: META[v]={'section':str(r[0] or '').strip(),'name':str(r[2] or v).strip(),
                       'def':str(r[3] or '').strip(),'src':str(r[4] or '').strip(),
                       'years':str(r[5] or '').strip()}

ws=wb['CHP_all_data']
rows=list(ws.iter_rows(values_only=True))
hdr=[str(c) if c is not None else '' for c in rows[1]]
data=[r for r in rows[2:] if isinstance(r[0],(int,float))]

GROUPED={
  'Race_Asian':'Asian','Race_Black':'Black','Race_Latino':'Latino',
  'Race_White':'White','Race_Other':'Other',
  'Age0to17':'under 18','Age18to24':'18 to 24','Age25to44':'25 to 44',
  'Age45to64':'45 to 64','Age65plus':'65 and over',
  'Edu_Did_Not_Complete_HS':'did not finish high school',
  'Edu_HSGrad_Some_College':'high school or some college',
  'Edu_College_Degree_And_Higher':'college degree or higher',
}
SKIP={'ID','lower_95CL','upper_95CL','NYC_Comparison_Pvalues','Name','Borough',''}
cols=[]
for i,h in enumerate(hdr):
    if h in SKIP: continue
    if h.startswith('Rank_') or h.startswith('Premature_Mort_Cause'): continue
    cds=[r for r in data if r[0]>=101]
    n=sum(1 for r in cds if isinstance(r[i],(int,float)))
    if n>=45: cols.append((i,h,n))

# how each variable should read
UNIT={'Overall_Pop':'people','Life_Expectancy':'years','Ratio_Bodega_Supermarket':'bodegas per supermarket',
      'Farmers_Markets':'markets','Premature_Mort_Number':'deaths','HIV_Diagnoses':'per 100,000',
      'HepC_Reports':'per 100,000','Infant_Mort':'per 1,000 live births',
      'Premature_Mort_Rate':'per 100,000','Assault_Hosp':'per 100,000','Psych_Hosp':'per 100,000',
      'Avoidable_Adult_Hosp':'per 100,000','Falls_Hosp':'per 100,000','Pedestrian_Hosp':'per 100,000',
      'Child_Asthma':'per 10,000','Air_Pollution':'mcg per cubic metre','Jail_Incarceration':'per 100,000',
      'Avertable_Death':'percent','Bike_Coverage':'percent','School_Absent':'percent'}

vals={}; meta={}
names={}; boro={}
for i,h,n in cols:
    series={}
    for r in data:
        v=r[i]
        if not isinstance(v,(int,float)): continue
        gid=int(r[0])
        key = 'city' if gid==0 else ('b'+str(gid) if gid<10 else str(gid))
        series[key]=round(float(v),3)
    if len(series)<45: continue
    m=META.get(h,{})
    # several variables share one metadata row, so Race_White and Race_Black
    # both came back labelled "Race/Ethnicity". Fall back to the column name.
    lab=m.get('name') or h.replace('_',' ')
    if h in GROUPED:
        lab=lab.rstrip()+': '+GROUPED[h]
    vals[h]=series
    meta[h]={'label':lab,
             'section':m.get('section') or 'Other',
             'def':m.get('def',''),'src':m.get('src',''),'years':m.get('years',''),
             'unit':UNIT.get(h,'percent')}

for r in data:
    gid=int(r[0])
    if gid>=101: names[str(gid)]=str(r[2]).strip()
    elif 0<gid<10: boro['b'+str(gid)]=str(r[2]).strip()

out={'source':'NYC Department of Health and Mental Hygiene, 2026 Community Health Profiles public use dataset',
     'url':'https://www.nyc.gov/site/doh/data/data-publications/profiles.page',
     'names':names,'boroughs':boro,'meta':meta,'values':vals}
p=R+'data/health/chp.json'
json.dump(out,open(p,'w'),separators=(',',':'))
print('variables:',len(vals),'| bytes:',os.path.getsize(p))
import collections
print('sections:',collections.Counter(m['section'] for m in meta.values()))
